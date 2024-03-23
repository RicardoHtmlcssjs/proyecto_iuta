from flask import session
import openpyxl
from openpyxl.styles.borders import Border, Side
from datetime import datetime
from io import BytesIO
from model.config import Db

class Acciones():
	# verfificar si un campo esta vacio
	def val_cam_vacio(self, var):
		text = len(var)
		if text == 0:
			return False
		else:
			return True 
	# verificar si la session existe
	def session(self):
		if 'id_usu_log' in session:
			log = True
		else:
			log = False
		return log
	#destruir sessiones creadas para cerrar session
	def cerrar_session(self):
		session.clear()
	# crear excel
	def reporte_clientes(self):
		# Crear un nuevo libro de trabajo
		wb = openpyxl.Workbook()

		# Crear una hoja de cálculo
		sheet = wb.active

		# Escribir datos en las celdas
		sheet.merge_cells('A1:G1')
		sheet['A1'] = 'LISTADO DE CLIENTES'
		sheet['A1'].font = openpyxl.styles.Font(size=20, color='FFFFFF', bold=True)
		sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
		sheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='center')
		sheet['B2'].alignment = openpyxl.styles.Alignment(horizontal='center')
		sheet['C2'].alignment = openpyxl.styles.Alignment(horizontal='center')
		sheet['D2'].alignment = openpyxl.styles.Alignment(horizontal='center')
		sheet['E2'].alignment = openpyxl.styles.Alignment(horizontal='center')
		sheet['F2'].alignment = openpyxl.styles.Alignment(horizontal='center')
		sheet['G2'].alignment = openpyxl.styles.Alignment(horizontal='center')
		sheet['A1'].fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='00B631')

		sheet['A2'] = 'N.'
		sheet['B2'] = 'CEDULA'
		sheet['C2'] = 'NOMBRE'
		sheet['D2'] = 'TELEFONO'
		sheet['E2'] = 'CORREO'
		sheet['F2'] = 'FEC. INSCIPCION'
		sheet['G2'] = 'FEC. VENCIMIENTO'
		sheet['A2'].font = openpyxl.styles.Font(bold=True)
		sheet['B2'].font = openpyxl.styles.Font(bold=True)
		sheet['C2'].font = openpyxl.styles.Font(bold=True)
		sheet['D2'].font = openpyxl.styles.Font(bold=True)
		sheet['E2'].font = openpyxl.styles.Font(bold=True)
		sheet['F2'].font = openpyxl.styles.Font(bold=True)
		sheet['G2'].font = openpyxl.styles.Font(bold=True)

		# ancho columnas
		sheet.column_dimensions['B'].width = 20
		sheet.column_dimensions['C'].width = 25
		sheet.column_dimensions['D'].width = 20
		sheet.column_dimensions['E'].width = 25
		sheet.column_dimensions['F'].width = 25
		sheet.column_dimensions['G'].width = 20

		datos_db = Db().fetchall("SELECT usuarios.nombre, usuarios.apellido, usuarios.cedula, usuarios.telefono, usuarios.correo, clientes.fecha_reg, clientes.fec_venci  FROM usuarios INNER JOIN  clientes ON clientes.fk_usuario = usuarios.id_usuario  WHERE fk_role = 2")
		col = 3
		num = 1

		fecha_actual = datetime.now()
		# Formatear la fecha
		fecha_formateada = fecha_actual.strftime("%d-%m-%Y")
		for row in datos_db:
			fecha_ins = str(row[5])
			fecha_ven = str(row[6])
			nom_completo = row[0] + " " + row[1]
			sheet[f"A{col}"] = str(num)
			sheet[f"B{col}"] = row[2]
			sheet[f"C{col}"] = nom_completo
			sheet[f"D{col}"] = row[3]
			sheet[f"E{col}"] = row[4]
			sheet[f"F{col}"] = fecha_ins[0:10]
			sheet[f"G{col}"] = fecha_ven[0:10]
			# texto a la izquierda
			sheet[f"A{col}"].alignment = openpyxl.styles.Alignment(horizontal='left')
			sheet[f"B{col}"].alignment = openpyxl.styles.Alignment(horizontal='left')
			sheet[f"C{col}"].alignment = openpyxl.styles.Alignment(horizontal='left')
			sheet[f"D{col}"].alignment = openpyxl.styles.Alignment(horizontal='left')
			sheet[f"E{col}"].alignment = openpyxl.styles.Alignment(horizontal='left')
			sheet[f"F{col}"].alignment = openpyxl.styles.Alignment(horizontal='left')
			sheet[f"G{col}"].alignment = openpyxl.styles.Alignment(horizontal='left')
			col = col + 1
			num = num + 1

		# Agregar una imagen
		# img = openpyxl.drawing.Image('imagen.png')
		# sheet.add_image(img, 'C1')

		# Guardar el archivo
		wb.save(f"static/reportes/{str(fecha_formateada)}_mi_archivo.xlsx")
		output = BytesIO()
		wb.save(output)
		

		return output